from __future__ import annotations

"""
PPM Productivity App (V2) - AMC Reference based selection + improved UI

Changes vs previous version:
1) Uses the attached "PPM App Data.xlsx" workbook (Project Planning Updated / Team / users).
2) Users choose Project Owner -> AMC Reference (Project Name is hidden in UI for security).
   Project Name is still stored in the backend Inputs sheet.
3) Improved UI/UX with Instacool-inspired styling, cleaner layout, and a form-based submission flow.

How data works:
- Source sheets (read-only):
  - "Project Planning Updated"  -> project master
  - "Team"                     -> technician / helper names
  - "users"                    -> phone-based login + roles (admin/user)
- Target sheet (read/write):
  - "Inputs"                   -> technician daily submissions (auto-created if missing)
"""

import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence, Tuple

import pandas as pd
import streamlit as st


# -------------------------
# Data models
# -------------------------


@dataclass(frozen=True)
class Submission:
    submission_date: dt.date
    project_owner: str
    amc_reference: str
    project_name_backend: str
    emirate: str
    ppm_no: str
    ppm_date: str
    indoors_completed: int
    vrf_completed: int
    dx_completed: int
    ahu_completed: int
    technicians: Sequence[str]
    helpers: Sequence[str]


@dataclass(frozen=True)
class RemainingCounts:
    indoors: int
    vrf: int
    dx: int
    ahu: int


class DataLoadError(RuntimeError):
    """Raised when the Excel workbook or its structure is invalid."""


# -------------------------
# Configuration
# -------------------------

WORKBOOK_NAME = "PPM App Data.xlsx"
THIS_DIR = Path(__file__).resolve().parent
DEFAULT_PATHS = [
    THIS_DIR / WORKBOOK_NAME,
    Path("/mnt/data") / WORKBOOK_NAME,  # for ChatGPT / container runs
]

SHEET_PROJECTS = "Project Planning Updated"
SHEET_TEAM = "Team"
SHEET_USERS = "users"
SHEET_INPUTS = "Inputs"

INPUT_COLUMNS = [
    "Date",
    "Project Owner",
    "AMC Reference",
    "Project Name (Backend)",
    "Emirate",
    "PPM #",
    "PPM Date",
    "Indoors Completed",
    "VRF OD Completed",
    "DX Outdoor Completed",
    "AHU Completed",
    "Technician name 1",
    "Technician name 2",
    "Technician name 3",
    "Helper name 1",
    "Helper name 2",
    "Helper name 3",
    "Submitted By (Phone)",
    "Submitted At",
]


# -------------------------
# Helpers
# -------------------------


def resolve_workbook_path() -> Path:
    for p in DEFAULT_PATHS:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"Could not find '{WORKBOOK_NAME}'. "
        f"Place it next to this script ({THIS_DIR}) or at /mnt/data/{WORKBOOK_NAME}."
    )


def _safe_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def _coerce_int(x) -> int:
    try:
        if pd.isna(x):
            return 0
        return int(float(x))
    except Exception:
        return 0


def app_header() -> None:
    """Page config + Instacool-ish CSS."""
    # "centered" behaves better on mobile; columns will stack naturally.
    st.set_page_config(page_title="PPM Productivity App", layout="centered")

    primary = "#00ADEF"
    accent = "#0090C6"
    bg = "#F6FBFE"
    card = "#FFFFFF"
    border = "#E6EEF5"
    muted = "#6B7280"

    st.markdown(
        f"""
        <style>
        .stApp {{
            background: {bg};
        }}
        /* Make main container a bit wider */
        .block-container {{
            padding-top: 1.2rem;
            padding-bottom: 2.5rem;
            max-width: 1200px;
        }}
        /* Card look for containers */
        div[data-testid="stVerticalBlockBorderWrapper"] {{
            background: {card};
            border: 1px solid {border};
            border-radius: 14px;
            padding: 14px 14px 6px 14px;
        }}
        /* Buttons */
        .stButton>button {{
            background: {primary};
            color: white;
            border-radius: 10px;
            border: 0px;
            padding: 0.6rem 1rem;
            font-weight: 600;
        }}
        .stButton>button:hover {{
            background: {accent};
            color: white;
        }}
        /* Labels */
        label, .stMarkdown, .stTextInput, .stSelectbox, .stNumberInput {{
            color: #111827;
        }}
        small, .stCaption, .st-emotion-cache-16idsys {{
            color: {muted};
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div style="display:flex; align-items:center; gap:12px; margin-bottom:8px;">
          <div style="width:10px; height:32px; background:#00ADEF; border-radius:8px;"></div>
          <div>
            <div style="font-size:24px; font-weight:800; line-height:1.1;">PPM Productivity</div>
            <div style="color:#6B7280; font-size:13px;">Daily technician submission (Owner → AMC Reference) • Project names are hidden</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# -------------------------
# Loading / saving
# -------------------------


def load_sheets(path: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load Projects, Team, Users, Inputs (Inputs auto-created if missing)."""
    try:
        xls = pd.ExcelFile(path)
    except Exception as exc:
        raise DataLoadError(f"Failed to open workbook: {path}") from exc

    for required in (SHEET_PROJECTS, SHEET_TEAM, SHEET_USERS):
        if required not in xls.sheet_names:
            raise DataLoadError(f"Missing required sheet '{required}' in workbook")

    projects = pd.read_excel(xls, sheet_name=SHEET_PROJECTS)
    team = pd.read_excel(xls, sheet_name=SHEET_TEAM)
    users = pd.read_excel(xls, sheet_name=SHEET_USERS)

    if SHEET_INPUTS in xls.sheet_names:
        inputs = pd.read_excel(xls, sheet_name=SHEET_INPUTS)
    else:
        inputs = pd.DataFrame(columns=INPUT_COLUMNS)

    # Normalize key columns
    for col in ["Project Owner", "AMC Reference", "Project Name", "Emirate", "PPM #", "PPM Date", "Month", "Year"]:
        if col in projects.columns:
            projects[col] = projects[col].map(_safe_str)

    if "Technician Name" in team.columns:
        team["Technician Name"] = team["Technician Name"].map(_safe_str)

    if not users.empty:
        for col in ["phone", "name", "role"]:
            if col in users.columns:
                users[col] = users[col].map(_safe_str)
    else:
        users = pd.DataFrame(columns=["phone", "name", "role"])

    # Ensure Inputs columns exist (upgrade safely)
    for c in INPUT_COLUMNS:
        if c not in inputs.columns:
            inputs[c] = ""

    return projects, team, users, inputs


def save_sheet_replace(path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    """Replace a single sheet in an existing workbook.

    NOTE: Kept only for rare admin maintenance tasks. Avoid using this for Inputs in production
    because replacing the sheet can accidentally wipe history if a read/load fails.
    """
    import openpyxl  # noqa: F401

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def _ensure_sheet_with_header(wb, sheet_name: str, columns: list[str]):
    """Ensure sheet exists and has the expected header row."""
    import openpyxl

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # If sheet is empty, create header
        if ws.max_row < 1:
            for j, col in enumerate(columns, start=1):
                ws.cell(row=1, column=j, value=col)
            return ws

        # Read header row
        header = [ws.cell(row=1, column=j).value for j in range(1, len(columns) + 1)]
        header = [str(h).strip() if h is not None else "" for h in header]

        if header != columns:
            # Do not overwrite - raise a clear error so user can fix workbook safely
            raise RuntimeError(
                f"Inputs sheet header mismatch. Expected: {columns} | Found: {header}. "
                "Please align the Inputs sheet columns to avoid data loss."
            )
        return ws

    # Create sheet with header
    ws = wb.create_sheet(title=sheet_name)
    for j, col in enumerate(columns, start=1):
        ws.cell(row=1, column=j, value=col)
    return ws


def append_rows_to_sheet(path: Path, sheet_name: str, df_new_rows: pd.DataFrame, columns: list[str]) -> None:
    """Append rows to an Excel sheet WITHOUT replacing it (prevents accidental wipes).

    This is much safer than if_sheet_exists='replace'. It appends new submissions to the end.
    """
    from openpyxl import load_workbook

    if df_new_rows.empty:
        return

    # Reindex to expected column order
    df_new_rows = df_new_rows.reindex(columns=columns)

    wb = load_workbook(path)
    ws = _ensure_sheet_with_header(wb, sheet_name, columns)

    start_row = ws.max_row + 1
    # Write each row
    for i, row in enumerate(df_new_rows.itertuples(index=False), start=0):
        for j, value in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j, value=value)

    wb.save(path)


def backup_submission_csv(path: Path, df_new_rows: pd.DataFrame) -> None:
    """Create a small CSV log backup next to the workbook.

    On Streamlit Community Cloud, the filesystem is still ephemeral, but this reduces the risk
    of accidental overwrites and helps local testing. For true persistence use Google Sheets/DB.
    """
    try:
        backup_dir = path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y%m%d")
        f = backup_dir / f"inputs_log_{stamp}.csv"
        # Append mode
        header = not f.exists()
        df_new_rows.to_csv(f, mode="a", index=False, header=header)
    except Exception:
        # Backups should never block a submission
        pass


def save_inputs_and_users(path: Path, inputs: pd.DataFrame, users: pd.DataFrame) -> None:
    """Legacy helper: avoid for Inputs; use append_rows_to_sheet for submissions."""
    save_sheet_replace(path, SHEET_USERS, users)

# -------------------------
# Business logic
# -------------------------


def owner_list(projects: pd.DataFrame) -> list[str]:
    if "Project Owner" not in projects.columns:
        return []
    owners = sorted([o for o in projects["Project Owner"].dropna().unique().tolist() if str(o).strip()])
    return owners


def amc_list_for_owner(projects: pd.DataFrame, owner: str) -> list[str]:
    df = projects.loc[projects["Project Owner"] == owner]
    if df.empty:
        return []
    amcs = sorted([a for a in df["AMC Reference"].dropna().unique().tolist() if str(a).strip()])
    return amcs


def project_backend_record(projects: pd.DataFrame, owner: str, amc_ref: str) -> pd.Series | None:
    """
    Return one representative row for a given owner + AMC reference.
    The dataset may have duplicates (e.g., one row per PPM). We pick the first non-empty.
    """
    df = projects.loc[(projects["Project Owner"] == owner) & (projects["AMC Reference"] == amc_ref)]
    if df.empty:
        return None
    # Prefer rows where "Project Name" exists
    df2 = df.copy()
    if "Project Name" in df2.columns:
        df2 = df2.sort_values(by=["Project Name"], na_position="last")
    return df2.iloc[0]


def ppm_options(projects: pd.DataFrame, owner: str, amc_ref: str) -> list[tuple[str, str]]:
    """
    Returns a list of (ppm_no, ppm_date) for owner+amc.
    """
    df = projects.loc[(projects["Project Owner"] == owner) & (projects["AMC Reference"] == amc_ref)]
    if df.empty:
        return []
    ppm_no_col = "PPM #"
    ppm_date_col = "PPM Date"
    if ppm_no_col not in df.columns or ppm_date_col not in df.columns:
        return []
    out = []
    for _, r in df[[ppm_no_col, ppm_date_col]].dropna(how="all").drop_duplicates().iterrows():
        out.append((_safe_str(r.get(ppm_no_col, "")), _safe_str(r.get(ppm_date_col, ""))))
    # Sort by ppm_no numeric if possible, otherwise as string
    def _key(t):
        n = t[0]
        try:
            return int(float(n))
        except Exception:
            return 999999
    out = sorted(out, key=_key)
    return out


def compute_remaining(
    projects: pd.DataFrame,
    inputs: pd.DataFrame,
    owner: str,
    amc_ref: str,
) -> RemainingCounts:
    rec = project_backend_record(projects, owner, amc_ref)
    if rec is None:
        return RemainingCounts(0, 0, 0, 0)

    total_indoor = _coerce_int(rec.get("Indoors Qty", 0))
    total_vrf = _coerce_int(rec.get("VRF OD Qty", 0))
    total_dx = _coerce_int(rec.get("DX Outdoor Qty", 0))
    total_ahu = _coerce_int(rec.get("AHU Qty", 0))

    existing = inputs.loc[
        (inputs["Project Owner"].map(_safe_str) == owner) & (inputs["AMC Reference"].map(_safe_str) == amc_ref)
    ]

    completed_indoor = _coerce_int(existing["Indoors Completed"].sum())
    completed_vrf = _coerce_int(existing["VRF OD Completed"].sum())
    completed_dx = _coerce_int(existing["DX Outdoor Completed"].sum())
    completed_ahu = _coerce_int(existing["AHU Completed"].sum())

    return RemainingCounts(
        indoors=max(total_indoor - completed_indoor, 0),
        vrf=max(total_vrf - completed_vrf, 0),
        dx=max(total_dx - completed_dx, 0),
        ahu=max(total_ahu - completed_ahu, 0),
    )


def append_submission(inputs: pd.DataFrame, sub: Submission, submitted_by_phone: str) -> pd.DataFrame:
    techs = list(sub.technicians)[:3] + [""] * 3
    helpers = list(sub.helpers)[:3] + [""] * 3

    row = {
        "Date": sub.submission_date,
        "Project Owner": sub.project_owner,
        "AMC Reference": sub.amc_reference,
        "Project Name (Backend)": sub.project_name_backend,
        "Emirate": sub.emirate,
        "PPM #": sub.ppm_no,
        "PPM Date": sub.ppm_date,
        "Indoors Completed": int(sub.indoors_completed),
        "VRF OD Completed": int(sub.vrf_completed),
        "DX Outdoor Completed": int(sub.dx_completed),
        "AHU Completed": int(sub.ahu_completed),
        "Technician name 1": techs[0],
        "Technician name 2": techs[1],
        "Technician name 3": techs[2],
        "Helper name 1": helpers[0],
        "Helper name 2": helpers[1],
        "Helper name 3": helpers[2],
        "Submitted By (Phone)": submitted_by_phone,
        "Submitted At": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    out = pd.concat([inputs, pd.DataFrame([row])], ignore_index=True)
    # Keep columns order stable
    out = out.reindex(columns=INPUT_COLUMNS)
    return out


# -------------------------
# Auth (phone based, using users sheet)
# -------------------------


def ensure_user(users: pd.DataFrame, phone: str) -> tuple[pd.DataFrame, bool]:
    """Ensure phone exists in users sheet; returns (updated_users, is_admin)."""
    phone = phone.strip()
    if phone == "":
        return users, False

    if "phone" not in users.columns:
        users = pd.DataFrame(columns=["phone", "name", "role"])

    if phone in users["phone"].astype(str).values:
        rec = users.loc[users["phone"].astype(str) == phone].iloc[0]
        is_admin = _safe_str(rec.get("role", "")).lower() == "admin"
        return users, is_admin

    new_row = pd.DataFrame([{"phone": phone, "name": "", "role": "user"}])
    users2 = pd.concat([users, new_row], ignore_index=True)
    return users2, False


def login_box(users: pd.DataFrame) -> tuple[bool, bool, str, pd.DataFrame]:
    """Returns (authenticated, is_admin, phone, users_df_updated)."""
    if "auth_done" not in st.session_state:
        st.session_state.auth_done = False
        st.session_state.phone = ""
        st.session_state.is_admin = False

    with st.container():
        st.subheader("Sign in")
        st.caption("Enter your mobile number to continue. (No OTP in this version)")
        colA, colB = st.columns([2, 1])
        with colA:
            phone = st.text_input("Mobile number", value=st.session_state.phone, placeholder="e.g., 05XXXXXXXX")
        with colB:
            st.write("")
            clicked = st.button("Login", use_container_width=True)

        if clicked:
            phone_clean = phone.strip()
            if not phone_clean:
                st.error("Please enter your mobile number.")
                return False, False, "", users

            users2, is_admin = ensure_user(users, phone_clean)
            st.session_state.auth_done = True
            st.session_state.phone = phone_clean
            st.session_state.is_admin = is_admin
            return True, is_admin, phone_clean, users2

        if st.session_state.auth_done:
            return True, st.session_state.is_admin, st.session_state.phone, users

    return False, False, "", users


# -------------------------
# UI
# -------------------------


def main() -> None:
    app_header()

    try:
        path = resolve_workbook_path()
        projects, team, users, inputs = load_sheets(path)
    except Exception as exc:
        st.error(f"Cannot load workbook: {exc}")
        st.stop()

    # Login
    authed, is_admin, phone, users_updated = login_box(users)
    if not authed:
        st.stop()

    # Persist newly created user (if any)
    if len(users_updated) != len(users) or not users_updated.equals(users):
        try:
            save_sheet_replace(path, SHEET_USERS, users_updated)
            users = users_updated
        except Exception:
            # don't block app if user save fails; but show warning
            st.warning("Could not update 'users' sheet. New users may not be saved.")

    st.markdown("---")
    st.write(f"Logged in as **{phone}** • Role: **{'Admin' if is_admin else 'User'}**")

    tabs = ["Submit Entry"]
    if is_admin:
        tabs.append("Summary & Export")
    tab = st.tabs(tabs)

    # -------- Submit tab --------
    with tab[0]:
        # Top selection row
        owners = owner_list(projects)
        if not owners:
            st.error("No project owners found in the project sheet.")
            st.stop()

        sel1, sel2, sel3 = st.columns([2, 2, 2])
        with sel1:
            owner = st.selectbox("Project Owner", owners, key="owner_select")
        with sel2:
            amcs = amc_list_for_owner(projects, owner)
            if not amcs:
                st.warning("No AMC references found for this owner.")
                st.stop()
            amc_ref = st.selectbox("AMC Reference", amcs, key="amc_select")
        with sel3:
            ppm_pairs = ppm_options(projects, owner, amc_ref)
            if ppm_pairs:
                ppm_labels = [f"PPM {p[0]}  •  {p[1]}" if p[0] or p[1] else "PPM" for p in ppm_pairs]
                ppm_choice = st.selectbox("PPM", ppm_labels, index=0)
                ppm_no, ppm_date = ppm_pairs[ppm_labels.index(ppm_choice)]
            else:
                ppm_no, ppm_date = "", ""
                st.selectbox("PPM", ["(not available)"], disabled=True)

        rec = project_backend_record(projects, owner, amc_ref)
        if rec is None:
            st.error("Could not find a matching project record for this Owner + AMC Reference.")
            st.stop()

        emirate = _safe_str(rec.get("Emirate", ""))
        project_name_backend = _safe_str(rec.get("Project Name", ""))

        # Remaining + totals
        # IMPORTANT: Per security requirement, technicians should NOT see unit totals/remaining or other users' submissions.
        remaining = compute_remaining(projects, inputs, owner, amc_ref)
        total_indoor = _coerce_int(rec.get("Indoors Qty", 0))
        total_vrf = _coerce_int(rec.get("VRF OD Qty", 0))
        total_dx = _coerce_int(rec.get("DX Outdoor Qty", 0))
        total_ahu = _coerce_int(rec.get("AHU Qty", 0))

        if is_admin:
            m1, m2, m3, m4, m5 = st.columns([1.1, 1, 1, 1, 1])
            with m1:
                st.metric("Emirate", emirate if emirate else "—")
            with m2:
                st.metric("Indoors Remaining", f"{remaining.indoors}/{total_indoor}")
            with m3:
                st.metric("VRF OD Remaining", f"{remaining.vrf}/{total_vrf}")
            with m4:
                st.metric("DX Remaining", f"{remaining.dx}/{total_dx}")
            with m5:
                st.metric("AHU Remaining", f"{remaining.ahu}/{total_ahu}")
            st.caption(
                "Project Name is hidden in the dropdown for security. It is still stored in the backend Inputs sheet."
            )
        else:
            # Minimal info for technicians
            st.info("Project is selected. Project Name and unit progress are hidden for security.")

        # Submission form
        tech_names = sorted([n for n in team.get("Technician Name", pd.Series(dtype=str)).dropna().unique().tolist() if str(n).strip()])

        with st.form("submit_form", clear_on_submit=True):
            left, right = st.columns([2, 1])
            with left:
                sub_date = st.date_input("Submission date", value=dt.date.today())
            with right:
                st.write("")
                st.write("")

            st.subheader("Completed Units (Today)")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                # Keep backend validation, but do not display remaining counts to technicians.
                indoors_completed = st.number_input(
                    "Indoors",
                    min_value=0,
                    max_value=100_000,
                    step=1,
                    value=0,
                )
            with c2:
                vrf_completed = st.number_input(
                    "VRF OD",
                    min_value=0,
                    max_value=100_000,
                    step=1,
                    value=0,
                )
            with c3:
                dx_completed = st.number_input(
                    "DX Outdoor",
                    min_value=0,
                    max_value=100_000,
                    step=1,
                    value=0,
                )
            with c4:
                ahu_completed = st.number_input(
                    "AHU",
                    min_value=0,
                    max_value=100_000,
                    step=1,
                    value=0,
                )

            st.subheader("Team")
            tcol, hcol = st.columns(2)
            with tcol:
                tech_selected = st.multiselect("Technician(s)", tech_names, max_selections=3)
            with hcol:
                helper_selected = st.multiselect("Helper(s)", tech_names, max_selections=3)

            submitted = st.form_submit_button("Submit", use_container_width=True)

        if submitted:
            # Optional: warn admins if entered quantities exceed remaining counts (submission is still allowed).
            # Note: quantities may exceed remaining counts; submission is still allowed.
            if is_admin:
                over = []
                if int(indoors_completed) > int(remaining.indoors):
                    over.append('Indoors')
                if int(vrf_completed) > int(remaining.vrf):
                    over.append('VRF OD')
                if int(dx_completed) > int(remaining.dx):
                    over.append('DX Outdoor')
                if int(ahu_completed) > int(remaining.ahu):
                    over.append('AHU')
                if over:
                    st.warning('Entered quantities exceed remaining for this project (admin can still save): ' + ', '.join(over))
            sub = Submission(
                submission_date=sub_date,
                project_owner=owner,
                amc_reference=amc_ref,
                project_name_backend=project_name_backend,
                emirate=emirate,
                ppm_no=_safe_str(ppm_no),
                ppm_date=_safe_str(ppm_date),
                indoors_completed=int(indoors_completed),
                vrf_completed=int(vrf_completed),
                dx_completed=int(dx_completed),
                ahu_completed=int(ahu_completed),
                technicians=tech_selected,
                helpers=helper_selected,
            )
            # Build only the new row and append it to Excel (no overwrite)
            inputs2 = append_submission(inputs, sub, submitted_by_phone=phone)
            new_row_df = inputs2.tail(1).copy()
            try:
                append_rows_to_sheet(path, SHEET_INPUTS, new_row_df, INPUT_COLUMNS)
                backup_submission_csv(path, new_row_df)
            except Exception as exc:
                st.error(f"Failed to save your submission to Excel: {exc}")
            else:
                st.success("Saved! Your submission has been appended to the Inputs sheet (no overwrite).")
                inputs = inputs2  # refresh for remaining calculations

        # SECURITY: No recent submissions / unit visibility for technicians.
        # Admin can review data in the Admin tab.
        if is_admin and not inputs.empty:
            with st.expander("Recent Submissions (Admin only)", expanded=False):
                safe_cols = [
                    "Date",
                    "Project Owner",
                    "AMC Reference",
                    "Emirate",
                    "PPM #",
                    "Indoors Completed",
                    "VRF OD Completed",
                    "DX Outdoor Completed",
                    "AHU Completed",
                    "Technician name 1",
                    "Technician name 2",
                    "Technician name 3",
                    "Helper name 1",
                    "Helper name 2",
                    "Helper name 3",
                    "Submitted By (Phone)",
                    "Submitted At",
                ]
                safe_cols = [c for c in safe_cols if c in inputs.columns]
                st.dataframe(inputs[safe_cols].tail(20).astype(str), use_container_width=True)

    # -------- Admin tab --------
    if is_admin and len(tab) > 1:
        with tab[1]:
            st.subheader("Summary & Export")
            st.caption("Admin-only view. You can export the Inputs sheet, and optionally view project names.")

            show_project_name = st.checkbox("Show Project Name (Backend) in table", value=False)

            # Build summary per owner+amc
            # Deduplicate projects by owner+amc
            base = projects.dropna(subset=["Project Owner", "AMC Reference"]).copy()
            base["Project Owner"] = base["Project Owner"].map(_safe_str)
            base["AMC Reference"] = base["AMC Reference"].map(_safe_str)
            base = base.drop_duplicates(subset=["Project Owner", "AMC Reference"])

            summary_rows = []
            for _, r in base.iterrows():
                owner = _safe_str(r.get("Project Owner", ""))
                amc_ref = _safe_str(r.get("AMC Reference", ""))
                proj_name = _safe_str(r.get("Project Name", ""))
                emirate = _safe_str(r.get("Emirate", ""))

                total_indoor = _coerce_int(r.get("Indoors Qty", 0))
                total_vrf = _coerce_int(r.get("VRF OD Qty", 0))
                total_dx = _coerce_int(r.get("DX Outdoor Qty", 0))
                total_ahu = _coerce_int(r.get("AHU Qty", 0))

                prev = inputs.loc[
                    (inputs["Project Owner"].map(_safe_str) == owner) & (inputs["AMC Reference"].map(_safe_str) == amc_ref)
                ]
                comp_indoor = _coerce_int(prev["Indoors Completed"].sum())
                comp_vrf = _coerce_int(prev["VRF OD Completed"].sum())
                comp_dx = _coerce_int(prev["DX Outdoor Completed"].sum())
                comp_ahu = _coerce_int(prev["AHU Completed"].sum())

                def pct(c, t):
                    return 0.0 if t <= 0 else min(max(c / t, 0.0), 1.0)

                summary_rows.append(
                    {
                        "Project Owner": owner,
                        "AMC Reference": amc_ref,
                        "Emirate": emirate,
                        "Project Name (Backend)": proj_name,
                        "Indoors": f"{comp_indoor}/{total_indoor} ({pct(comp_indoor, total_indoor):.0%})",
                        "VRF OD": f"{comp_vrf}/{total_vrf} ({pct(comp_vrf, total_vrf):.0%})",
                        "DX Outdoor": f"{comp_dx}/{total_dx} ({pct(comp_dx, total_dx):.0%})",
                        "AHU": f"{comp_ahu}/{total_ahu} ({pct(comp_ahu, total_ahu):.0%})",
                    }
                )

            summary_df = pd.DataFrame(summary_rows)
            if not show_project_name and "Project Name (Backend)" in summary_df.columns:
                summary_df = summary_df.drop(columns=["Project Name (Backend)"])

            st.dataframe(summary_df, use_container_width=True, height=520)

            # Download Inputs
            cleaned = inputs.copy().drop_duplicates()
            csv_data = cleaned.to_csv(index=False)
            st.download_button(
                "Download Inputs as CSV",
                data=csv_data,
                file_name="PPM_inputs_export.csv",
                mime="text/csv",
                use_container_width=True,
            )


if __name__ == "__main__":
    main()
